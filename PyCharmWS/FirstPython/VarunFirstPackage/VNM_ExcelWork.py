import openpyxl
import csv

csv.register_dialect('pipes', delimiter='|')
csv.register_dialect('semiColon', delimiter=';')


ROOT_DIR = "D:/GeocodingProjects/Vietnam_VNM/16Sept_Runs/"
TECHNOBANK_LAT_LON_FILE = "csv/TechnoBank_Lat_Lon.csv"
CGGE_FILE = "csv/CGGE.txt"
COMBINED_FILE = "csv/VNM_Combined.txt"
ANALYSIS_FILE = "Analysis.xlsx"
CGGE_HEADER = "ID|IN_HNR|IN_MAINADDR|IN_AN4|IN_AN3|IN_AN2|IN_AN1|IN_PC1|IN_POSTADDR|IN_COUNTRY|CAND_HNR|CAND_MAINADDR|CAND_AN4|CAND_AN3|CAND_AN2|CAND_AN1|CAND_PC1|CAND_PC2|CAND_FSA|CAND_FLA|CAND_COUNTRY|CAND_RESULTCODE|CAND_X|CAND_Y"
LAT_LONG_FILE_HEADER="ID|Address|ExpDistrict|ExpProvince|ExpLat|ExpLong"
OUTPUT_FILE_HEADER = CGGE_HEADER+"|"+LAT_LONG_FILE_HEADER


def parseCSVFile(file):

    return None

def loadCGGE():
    cggeFileHandle = open(ROOT_DIR + CGGE_FILE, mode="r", encoding="UTF-8")
    dictOfData = csv.DictReader(cggeFileHandle, fieldnames=CGGE_HEADER.split(sep="|"), dialect="semiColon")
    cggeDictionary = {}
    for row in dictOfData:
        # cggeDictionary.append({'key': row['IN_MAINADDR'], 'row': row})
        cggeDictionary[row['IN_MAINADDR'].upper().replace(' ','')]=row
    return cggeDictionary

    return None



def loadLat_LonFile():
    latLonFileHandle = open(ROOT_DIR+TECHNOBANK_LAT_LON_FILE,mode="r",encoding="UTF-8")
    dictOfData = csv.DictReader(latLonFileHandle,fieldnames=LAT_LONG_FILE_HEADER.split(sep="|"),dialect="pipes")
    latLonDictionary = {}
    rowCount = 0
    for row in dictOfData:
        if (dictOfData.line_num == 1):
            continue
        latLonDictionary[row['Address'].replace(' ','')]=row
        rowCount = rowCount + 1
        # print(str(rowCount)+":"+row['Address'])

    return latLonDictionary


def loadLucene():
    wb = openpyxl.load_workbook(ROOT_DIR + ANALYSIS_FILE)
    print(type(wb))
    print(wb.get_sheet_names())
    luceneSheet = wb.get_sheet_by_name('Lucene')
    print(luceneSheet.max_row, luceneSheet.max_column)
    rows = luceneSheet.rows
    first_row = [cell.value for cell in next(rows)]
    print(first_row)
    dict = {}
    for row in rows:
        record={}
        for key, cell in zip(first_row, row):
            record[key]=cell.value
            if (key == 'InputAddress'):
                record[key] = str(record[key]).replace(", VIETNAM","").upper()

        # dict.append({'key':record['InputAddress'],'value':record})
        dict[record['InputAddress'].replace(' ','')]=record


    return dict






def testCSVWithHeader():
    latLonFileHandle = open(ROOT_DIR + TECHNOBANK_LAT_LON_FILE, mode="r", encoding="UTF-8")
    dictOfData = csv.DictReader(latLonFileHandle, fieldnames=LAT_LONG_FILE_HEADER.split(sep="|"), dialect="pipes")
    latLonDictionary = {}
    rowCount = 0
    for row in dictOfData:
        if(dictOfData.line_num==1):
            continue
        latLonDictionary[row['Address'].replace(' ', '')] = row
        rowCount = rowCount + 1
        # print(str(rowCount)+":"+row['Address'])

    return latLonDictionary



def testVarun():
    # testCSVWithHeader()





    return None


def writeDictionaryListToCSV(filePath,csvDictionaryList, fieldNamesList):

    with open(filePath, 'w', encoding="UTF-8", newline='') as outcsv:

        writer = csv.DictWriter(outcsv,fieldnames=fieldNamesList,dialect="pipes")
        writer.writeheader()
        writer.writerows(row for row in csvDictionaryList)
    outcsv.close()
    pass


def generateCombinedFile(latLonDict, cggeDict, luceneDictionary):
    combineDict = []
    found = 0
    print("keyCount in lat - lon dictionary: "+str(len(latLonDict.keys())))
    for key, value in latLonDict.items():
        combineRecord = {}
        inputAddress=key.upper()

        cggeRecord = cggeDict.get(inputAddress)
        luceneRecord = luceneDictionary.get(inputAddress)

        if((cggeRecord == None)  or (luceneRecord==None)):
            print("workign on : " + value['Address'])
            print("cggeRecord = "+str(cggeRecord) + ": : lucenceRecord = " + str(luceneRecord))
        else:
            found = found+1


        combineRecord['inputAddress']=value['Address']
        combineRecord['expAN2']=value['ExpDistrict']
        combineRecord['expAN1'] = value['ExpProvince']
        combineRecord['ExpLat'] = value['ExpLat']
        combineRecord['ExpLong'] = value['ExpLong']

        if(cggeRecord==None): cggeRecord={}
        combineRecord['cggeMainAddress']=cggeRecord.get('CAND_MAINADDR')
        combineRecord['cggeAN4'] =cggeRecord.get('CAND_AN4')
        combineRecord['cggeAN3'] =cggeRecord.get('CAND_AN3')
        combineRecord['cggeAN2'] =cggeRecord.get('CAND_AN2')
        combineRecord['cggeAN1'] =cggeRecord.get('CAND_AN1')
        combineRecord['cggeFSA'] =cggeRecord.get('CAND_FSA')
        combineRecord['cggeFLA'] =cggeRecord.get('CAND_FLA')
        combineRecord['cggeLatitude'] =cggeRecord.get('CAND_Y')
        combineRecord['cggeLongitude'] =cggeRecord.get('CAND_X')
        combineRecord['cggePrecision'] =cggeRecord.get('CAND_RESULTCODE')

        if(luceneRecord==None):luceneRecord={}
        combineRecord['luceneStreetName'] = luceneRecord.get('streetname')
        combineRecord['luceneAN4'] = luceneRecord.get('locality')
        combineRecord['luceneAN3'] = luceneRecord.get('town')
        combineRecord['luceneAN2'] = luceneRecord.get('district')
        combineRecord['luceneAN1'] = luceneRecord.get('state')
        combineRecord['luceneLatitude'] = luceneRecord.get('lattitude')
        combineRecord['luceneLongitude'] = luceneRecord.get('longitude')

        combineDict.append(combineRecord)





    print(combineDict, sep="\n")
    combinedFieldNameList  = ['inputAddress','expAN2','expAN1','ExpLat','ExpLong','cggeMainAddress','cggeAN4','cggeAN3','cggeAN2','cggeAN1','cggeFSA','cggeFLA','cggeLatitude','cggeLongitude','cggePrecision','luceneStreetName','luceneAN4','luceneAN3','luceneAN2','luceneAN1','luceneLatitude' ,'luceneLongitude',]
    writeDictionaryListToCSV(ROOT_DIR+COMBINED_FILE,combineDict,combinedFieldNameList)
    pass


def main():

    latLonDict = loadLat_LonFile()
    cggeDict = loadCGGE()
    luceneDictionary = loadLucene()

    generateCombinedFile(latLonDict,cggeDict,luceneDictionary)
    return None




main()

# lat = "384-386-388 Đồng Khởi, phường Tân Hiệp, Thành phố Biên Hòa, Đồng Nai"
# cgge = "384-386-388 Đồng Khởi, phường Tân Hiệp, Thành phố Biên Hòa, Đồng Nai"
# lucene = "384-386-388 ĐONG KHOI, TAN HIEP, THÀNH PHỐ BIÊN HÒA, ĐỒNG NAI, VIETNAM"
#
# print(lat.upper()==cgge.upper())